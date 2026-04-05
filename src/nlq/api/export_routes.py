"""
Excel export endpoints for AOS-NLQ.

GET /export/financial-statement — generates .xlsx from session-stored financial statement data.
"""

import io
import logging
from typing import Optional

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse, JSONResponse

from src.nlq.api.session import get_dashboard_session_store

logger = logging.getLogger(__name__)

router = APIRouter(tags=["export"])


def _get_openpyxl():
    """Import openpyxl or return None with error response."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        return openpyxl
    except ImportError:
        return None


@router.get("/export/financial-statement")
async def export_financial_statement(
    session_id: str = Query(..., description="Browser session ID"),
    format: str = Query(default="xlsx", description="Export format (only 'xlsx' supported)"),
):
    """Export the most recent financial statement as an Excel file."""
    if format != "xlsx":
        return JSONResponse(
            status_code=400,
            content={"error": f"Unsupported format '{format}'. Only 'xlsx' is supported."},
        )

    store = get_dashboard_session_store()
    fs_data = store.get_financial_statement(session_id)
    if not fs_data:
        return JSONResponse(
            status_code=404,
            content={
                "error": "No financial statement found for this session. "
                "Submit a P&L query first (e.g., 'show me the P&L').",
                "session_id": session_id,
            },
        )

    openpyxl = _get_openpyxl()
    if not openpyxl:
        logger.error("openpyxl not installed — cannot generate Excel export")
        return JSONResponse(
            status_code=500,
            content={"error": "Excel export requires openpyxl. Install with: pip install openpyxl"},
        )

    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    title = fs_data.get("title", "Income Statement")
    entity = fs_data.get("entity", "")
    periods = fs_data.get("periods", [])
    line_items = fs_data.get("line_items", [])
    unit = fs_data.get("unit", "millions")

    num_cols = 1 + len(periods)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    for i, _ in enumerate(periods):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    title_font = Font(name="Calibri", size=14, bold=True)
    subtitle_font = Font(name="Calibri", size=10, color="808080", italic=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    header_border = Border(bottom=Side(style="medium"))
    subtotal_font = Font(name="Calibri", size=10, bold=True)
    subtotal_border = Border(top=Side(style="thin"))
    currency_format = '#,##0.0;(#,##0.0);"-"'
    percent_format = '0.0%;(0.0%);"-"'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.cell(row=1, column=1, value=f"{title} — {entity}").font = title_font

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1, value=f"All amounts in ${unit[0].upper()} unless noted")
    sub_cell.font = subtitle_font
    sub_cell.alignment = Alignment(horizontal="right")

    ws.cell(row=3, column=1, value="").font = header_font
    ws.cell(row=3, column=1).border = header_border
    for i, period in enumerate(periods):
        cell = ws.cell(row=3, column=i + 2, value=period)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")
        cell.border = header_border

    for row_idx, item in enumerate(line_items, start=4):
        label = item.get("label", "")
        indent = item.get("indent", 0)
        fmt = item.get("format", "currency")
        is_subtotal = item.get("is_subtotal", False)
        item_values = item.get("values", {})

        label_text = ("    " * indent) + label
        label_cell = ws.cell(row=row_idx, column=1, value=label_text)
        if is_subtotal:
            label_cell.font = subtotal_font
            label_cell.border = subtotal_border

        for col_idx, period in enumerate(periods):
            val = item_values.get(period)
            cell = ws.cell(row=row_idx, column=col_idx + 2)
            cell.alignment = Alignment(horizontal="right")
            if val is not None:
                if fmt == "percent":
                    cell.value = val / 100.0
                    cell.number_format = percent_format
                else:
                    cell.value = val
                    cell.number_format = currency_format
            else:
                cell.value = "--"
            if is_subtotal:
                cell.font = subtotal_font
                cell.border = subtotal_border

    ws.freeze_panes = "B4"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="income_statement.xlsx"'},
    )
